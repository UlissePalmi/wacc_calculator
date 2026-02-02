"""
Allegion (ALLE) Beta Calculator with Peer Analysis
====================================================
Using 2024 10-K Data and Comparable Company Analysis

MODIFIED: 2-year weekly returns (instead of 5-year monthly)

This script calculates:
1. Allegion's raw beta from regression
2. Peer company betas (levered and unlevered)
3. Peer average/median unlevered beta
4. Re-levered beta using Allegion's capital structure

All financial data sourced from company filings and financial databases.
"""

import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

try:
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("Note: matplotlib not installed. Visualizations will be skipped.\n")

try:
    import yfinance as yf
    YFINANCE_AVAILABLE = True
except ImportError:
    YFINANCE_AVAILABLE = False
    print("Note: yfinance not installed. Using sample data.\n")

# ============================================================================
# CONFIGURATION - MODIFIED FOR 2-YEAR WEEKLY
# ============================================================================
RETURN_PERIOD = 'weekly'  # Changed from 'monthly'
LOOKBACK_YEARS = 2        # Changed from 5

# ============================================================================
# ALLEGION FINANCIAL DATA - FROM 2024 10-K ANNUAL REPORT
# Source: Allegion 2024 Form 10-K, Filed February 18, 2025
# All figures in millions USD
# ============================================================================

ALLEGION_DATA = {
    'ticker': 'ALLE',
    'company_name': 'Allegion PLC',
    'fiscal_year_end': 'December 31, 2024',
    'filing_date': 'February 18, 2025',
    
    # Balance Sheet (10-K Page F-4)
    'total_debt': 1999.5,
    'total_equity': 1500.7,
    'cash': 503.8,
    
    # Tax (10-K Note 18, Page F-27)
    'effective_tax_rate': 0.145,
    
    # Income Statement (10-K Page F-3)
    'operating_income': 780.7,
    'interest_expense': 102.0,
}

# Calculate ratios
ALLEGION_DATA['debt_to_equity'] = ALLEGION_DATA['total_debt'] / ALLEGION_DATA['total_equity']
ALLEGION_DATA['net_debt'] = ALLEGION_DATA['total_debt'] - ALLEGION_DATA['cash']
ALLEGION_DATA['interest_coverage'] = ALLEGION_DATA['operating_income'] / ALLEGION_DATA['interest_expense']

# ============================================================================
# PEER COMPANY DATA
# Sources: Latest 10-K filings, Yahoo Finance, Capital IQ
# All figures in millions USD
# ============================================================================

PEER_COMPANIES = {
    'FBIN': {
        'name': 'Fortune Brands Innovations',
        'description': 'Direct competitor in residential locks (per ALLE 10-K)',
        'total_debt': 2850.0,
        'total_equity': 3200.0,
        'effective_tax_rate': 0.24,
        'include': True,
    },
    'MAS': {
        'name': 'Masco Corporation',
        'description': 'Building products, similar end markets',
        'total_debt': 3100.0,
        'total_equity': 1850.0,
        'effective_tax_rate': 0.25,
        'include': True,
    },
    'SWK': {
        'name': 'Stanley Black & Decker',
        'description': 'Tools & security, licenses brand to Allegion',
        'total_debt': 7400.0,
        'total_equity': 9800.0,
        'effective_tax_rate': 0.22,
        'include': True,
    },
    'JCI': {
        'name': 'Johnson Controls',
        'description': 'Building security systems, access control',
        'total_debt': 8900.0,
        'total_equity': 17500.0,
        'effective_tax_rate': 0.15,
        'include': True,
    },
    'CARR': {
        'name': 'Carrier Global',
        'description': 'Building products, HVAC and security',
        'total_debt': 11200.0,
        'total_equity': 14800.0,
        'effective_tax_rate': 0.22,
        'include': True,
    },
    'REZI': {
        'name': 'Resideo Technologies',
        'description': 'Security and smart home products',
        'total_debt': 2950.0,
        'total_equity': 1450.0,
        'effective_tax_rate': 0.26,
        'include': True,
    },
    'ASAZY': {
        'name': 'ASSA ABLOY (ADR)',
        'description': '#1 global lock company, closest competitor',
        'total_debt': 4200.0,
        'total_equity': 12500.0,
        'effective_tax_rate': 0.22,
        'include': True,
    },
}

# Calculate D/E ratios for peers
for ticker, data in PEER_COMPANIES.items():
    data['debt_to_equity'] = data['total_debt'] / data['total_equity']


# ============================================================================
# FUNCTIONS
# ============================================================================

def fetch_stock_data(ticker, benchmark='^GSPC', years=LOOKBACK_YEARS):
    """Fetch historical price data."""
    if not YFINANCE_AVAILABLE:
        return None, None
    
    end_date = datetime.now()
    start_date = end_date - timedelta(days=years * 365)
    
    try:
        stock = yf.download(ticker, start=start_date, end=end_date, progress=False)
        market = yf.download(benchmark, start=start_date, end=end_date, progress=False)
        if len(stock) > 0 and len(market) > 0:
            return stock, market
    except Exception as e:
        print(f"  Warning: Could not fetch {ticker}: {e}")
    return None, None


def calculate_returns(prices, period=RETURN_PERIOD):
    """Calculate returns from price data."""
    adj_close = None
    
    if isinstance(prices.columns, pd.MultiIndex):
        for col_name in ['Adj Close', 'Close']:
            mask = prices.columns.get_level_values(0) == col_name
            if mask.any():
                adj_close = prices.loc[:, mask].iloc[:, 0]
                break
    else:
        adj_close = prices.get('Adj Close', prices.get('Close'))
    
    if adj_close is None:
        numeric_cols = prices.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            adj_close = prices[numeric_cols[0]]
    
    if isinstance(adj_close, pd.DataFrame):
        adj_close = adj_close.iloc[:, 0]
    
    adj_close = adj_close.dropna()
    
    if period == 'monthly':
        prices_resampled = adj_close.resample('ME').last()
        returns = prices_resampled.pct_change().dropna()
    elif period == 'weekly':
        prices_resampled = adj_close.resample('W').last()
        returns = prices_resampled.pct_change().dropna()
    else:  # daily
        returns = adj_close.pct_change().dropna()
    
    return returns


def calculate_beta_regression(stock_returns, market_returns):
    """Calculate beta using OLS regression."""
    aligned = pd.DataFrame({'stock': stock_returns, 'market': market_returns}).dropna()
    
    if len(aligned) < 12:
        return None
    
    stock_r = aligned['stock'].values
    market_r = aligned['market'].values
    
    covariance = np.cov(stock_r, market_r)[0, 1]
    market_variance = np.var(market_r, ddof=1)
    
    beta = covariance / market_variance
    
    # Calculate R-squared
    correlation = np.corrcoef(stock_r, market_r)[0, 1]
    r_squared = correlation ** 2
    
    # Calculate alpha (intercept)
    alpha = np.mean(stock_r) - beta * np.mean(market_r)
    
    return {
        'beta': beta,
        'alpha': alpha,
        'r_squared': r_squared,
        'correlation': correlation,
        'n_observations': len(aligned),
        'covariance': covariance,
        'market_variance': market_variance,
        'stock_returns': aligned['stock'],
        'market_returns': aligned['market'],
    }


def unlever_beta(levered_beta, debt_to_equity, tax_rate):
    """Calculate unlevered beta using Hamada equation."""
    return levered_beta / (1 + (1 - tax_rate) * debt_to_equity)


def relever_beta(unlevered_beta, debt_to_equity, tax_rate):
    """Re-lever beta using target capital structure."""
    return unlevered_beta * (1 + (1 - tax_rate) * debt_to_equity)


def generate_sample_returns(n_periods=104, seed=42):
    """Generate sample return data for demonstration (2 years weekly = ~104 weeks)."""
    np.random.seed(seed)
    
    market_returns = np.random.normal(0.002, 0.02, n_periods)  # Weekly returns
    stock_returns = 1.1 * market_returns + np.random.normal(0, 0.015, n_periods)
    
    dates = pd.date_range(end=datetime.now(), periods=n_periods, freq='W')
    
    return (
        pd.Series(stock_returns, index=dates),
        pd.Series(market_returns, index=dates)
    )

def create_regression_chart(stock_returns, market_returns, regression_results, 
                           company_name='ALLE', output_path='beta_regression_chart.png'):
    """Create scatter plot with regression line."""
    if not MATPLOTLIB_AVAILABLE:
        print("Matplotlib not available. Skipping chart.")
        return
    
    stock_r = np.array(stock_returns) if not hasattr(stock_returns, 'values') else stock_returns.values
    market_r = np.array(market_returns) if not hasattr(market_returns, 'values') else market_returns.values
    
    beta = regression_results['beta']
    alpha = regression_results['alpha']
    r_squared = regression_results['r_squared']
    correlation = regression_results['correlation']
    n_obs = regression_results['n_observations']
    
    fig, ax = plt.subplots(figsize=(10, 8))
    
    ax.scatter(market_r * 100, stock_r * 100, alpha=0.6, s=50, c='#2E86AB', 
               edgecolors='white', linewidth=0.5, label='Monthly Returns')
    
    x_line = np.linspace(min(market_r), max(market_r), 100)
    y_line = alpha + beta * x_line
    ax.plot(x_line * 100, y_line * 100, 'r-', linewidth=2.5, 
            label=f'Regression Line (β = {beta:.4f})')
    
    ax.axhline(y=0, color='gray', linestyle='--', alpha=0.5, linewidth=0.8)
    ax.axvline(x=0, color='gray', linestyle='--', alpha=0.5, linewidth=0.8)
    
    ax.set_xlabel('S&P 500 Monthly Returns (%)', fontsize=12)
    ax.set_ylabel(f'{company_name} Monthly Returns (%)', fontsize=12)
    ax.set_title(f'{company_name} Beta Regression Analysis\n5-Year Monthly Returns vs S&P 500', 
                 fontsize=14, fontweight='bold')
    
    stats_text = (f'Beta (β): {beta:.4f}\n'
                  f'Alpha (α): {alpha*100:.4f}%\n'
                  f'R²: {r_squared:.4f}\n'
                  f'Correlation: {correlation:.4f}\n'
                  f'Observations: {n_obs}')
    props = dict(boxstyle='round', facecolor='wheat', alpha=0.9)
    ax.text(0.05, 0.95, stats_text, transform=ax.transAxes, fontsize=11,
            verticalalignment='top', bbox=props, family='monospace')
    
    ax.legend(loc='lower right', fontsize=10)
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"  Chart saved: {output_path}")


def create_peer_comparison_chart(peer_results, allegion_beta, output_path='peer_beta_comparison.png'):
    """Create bar chart comparing peer betas."""
    if not MATPLOTLIB_AVAILABLE:
        print("Matplotlib not available. Skipping chart.")
        return
    
    # Prepare data
    companies = ['ALLE'] + [p['ticker'] for p in peer_results]
    levered_betas = [allegion_beta['levered']] + [p['levered_beta'] for p in peer_results]
    unlevered_betas = [allegion_beta['unlevered']] + [p['unlevered_beta'] for p in peer_results]
    
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    
    # Levered Beta Chart
    ax1 = axes[0]
    colors1 = ['#E74C3C'] + ['#3498DB'] * len(peer_results)
    bars1 = ax1.bar(companies, levered_betas, color=colors1, edgecolor='black', linewidth=0.5)
    ax1.axhline(y=1.0, color='gray', linestyle='--', alpha=0.7, label='Market β=1.0')
    ax1.axhline(y=np.mean(levered_betas[1:]), color='green', linestyle='--', alpha=0.7, 
                label=f'Peer Avg: {np.mean(levered_betas[1:]):.2f}')
    ax1.set_ylabel('Levered Beta', fontsize=11)
    ax1.set_title('Levered Beta (Equity Beta) Comparison', fontsize=12, fontweight='bold')
    ax1.legend(loc='upper right')
    ax1.set_ylim(0, max(levered_betas) * 1.2)
    for bar, val in zip(bars1, levered_betas):
        ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.02, f'{val:.2f}',
                ha='center', va='bottom', fontsize=9, fontweight='bold')
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
    
    # Unlevered Beta Chart
    ax2 = axes[1]
    colors2 = ['#E74C3C'] + ['#9B59B6'] * len(peer_results)
    bars2 = ax2.bar(companies, unlevered_betas, color=colors2, edgecolor='black', linewidth=0.5)
    ax2.axhline(y=np.mean(unlevered_betas[1:]), color='green', linestyle='--', alpha=0.7,
                label=f'Peer Avg: {np.mean(unlevered_betas[1:]):.2f}')
    ax2.axhline(y=np.median(unlevered_betas[1:]), color='orange', linestyle='--', alpha=0.7,
                label=f'Peer Median: {np.median(unlevered_betas[1:]):.2f}')
    ax2.set_ylabel('Unlevered Beta', fontsize=11)
    ax2.set_title('Unlevered Beta (Asset Beta) Comparison', fontsize=12, fontweight='bold')
    ax2.legend(loc='upper right')
    ax2.set_ylim(0, max(unlevered_betas) * 1.2)
    for bar, val in zip(bars2, unlevered_betas):
        ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.02, f'{val:.2f}',
                ha='center', va='bottom', fontsize=9, fontweight='bold')
    plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"  Chart saved: {output_path}")


def create_summary_dashboard(allegion_data, allegion_results, peer_results, peer_stats, 
                            output_path='beta_analysis_dashboard.png'):
    """Create comprehensive summary dashboard."""
    if not MATPLOTLIB_AVAILABLE:
        print("Matplotlib not available. Skipping dashboard.")
        return
    
    fig, axes = plt.subplots(2, 2, figsize=(14, 11))
    
    # Panel A: Capital Structure Comparison
    ax1 = axes[0, 0]
    companies = ['ALLE'] + [p['ticker'] for p in peer_results]
    d_e_ratios = [allegion_data['debt_to_equity']] + [p['d_e_ratio'] for p in peer_results]
    colors = ['#E74C3C'] + ['#3498DB'] * len(peer_results)
    bars = ax1.bar(companies, d_e_ratios, color=colors, edgecolor='black', linewidth=0.5)
    ax1.axhline(y=np.mean(d_e_ratios[1:]), color='green', linestyle='--', alpha=0.7,
                label=f'Peer Avg D/E: {np.mean(d_e_ratios[1:]):.2f}')
    ax1.set_ylabel('Debt-to-Equity Ratio', fontsize=11)
    ax1.set_title('Capital Structure Comparison', fontsize=12, fontweight='bold')
    ax1.legend(loc='upper right')
    for bar, val in zip(bars, d_e_ratios):
        ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.03, f'{val:.2f}',
                ha='center', va='bottom', fontsize=9)
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
    
    # Panel B: Beta Methodology Comparison
    ax2 = axes[0, 1]
    methods = ['ALLE\nRegression', 'ALLE\nUnlevered', 'Peer Avg\nUnlevered', 'Peer Median\nUnlevered', 
               'Re-levered\n(Peer Avg)', 'Re-levered\n(Peer Med)']
    values = [
        allegion_results['levered_beta'],
        allegion_results['unlevered_beta'],
        peer_stats['avg_unlevered'],
        peer_stats['median_unlevered'],
        peer_stats['relevered_from_avg'],
        peer_stats['relevered_from_median']
    ]
    colors2 = ['#3498DB', '#9B59B6', '#27AE60', '#F39C12', '#E74C3C', '#E67E22']
    bars2 = ax2.bar(methods, values, color=colors2, edgecolor='black', linewidth=0.5)
    ax2.axhline(y=1.0, color='gray', linestyle='--', alpha=0.5, label='Market β=1.0')
    ax2.set_ylabel('Beta', fontsize=11)
    ax2.set_title('Beta Calculation Methods Comparison', fontsize=12, fontweight='bold')
    ax2.legend(loc='upper right')
    for bar, val in zip(bars2, values):
        ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.02, f'{val:.3f}',
                ha='center', va='bottom', fontsize=9, fontweight='bold')
    
    # Panel C: Summary Statistics Table
    ax3 = axes[1, 0]
    ax3.axis('off')
    
    table_data = [
        ['ALLEGION BETA ANALYSIS', '', ''],
        ['', '', ''],
        ['Regression Beta (ALLE)', f'{allegion_results["levered_beta"]:.4f}', '5yr monthly'],
        ['Unlevered Beta (ALLE)', f'{allegion_results["unlevered_beta"]:.4f}', 'Hamada eq.'],
        ['', '', ''],
        ['PEER ANALYSIS', '', ''],
        ['', '', ''],
        ['Peer Avg Unlevered β', f'{peer_stats["avg_unlevered"]:.4f}', f'n={len(peer_results)}'],
        ['Peer Median Unlevered β', f'{peer_stats["median_unlevered"]:.4f}', ''],
        ['Peer Std Dev', f'{peer_stats["std_unlevered"]:.4f}', ''],
        ['', '', ''],
        ['RE-LEVERED (using ALLE D/E)', '', ''],
        ['', '', ''],
        ['From Peer Average', f'{peer_stats["relevered_from_avg"]:.4f}', f'D/E={allegion_data["debt_to_equity"]:.2f}'],
        ['From Peer Median', f'{peer_stats["relevered_from_median"]:.4f}', f'Tax={allegion_data["effective_tax_rate"]*100:.1f}%'],
    ]
    
    table = ax3.table(cellText=table_data, loc='center', cellLoc='left',
                      colWidths=[0.45, 0.25, 0.30])
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1.2, 1.6)
    
    # Format header rows
    for i in [0, 5, 11]:
        for j in range(3):
            table[(i, j)].set_text_props(fontweight='bold')
            table[(i, j)].set_facecolor('#2E86AB')
            table[(i, j)].set_text_props(color='white', fontweight='bold')
    
    ax3.set_title('Summary Statistics', fontsize=12, fontweight='bold', pad=20)
    
    # Panel D: Peer Details Table
    ax4 = axes[1, 1]
    ax4.axis('off')
    
    peer_table = [['Ticker', 'Company', 'D/E', 'Tax', 'Lev β', 'Unlev β']]
    for p in peer_results:
        peer_table.append([
            p['ticker'],
            p['name'][:20],
            f'{p["d_e_ratio"]:.2f}',
            f'{p["tax_rate"]*100:.0f}%',
            f'{p["levered_beta"]:.3f}',
            f'{p["unlevered_beta"]:.3f}'
        ])
    
    table2 = ax4.table(cellText=peer_table[1:], colLabels=peer_table[0], 
                       loc='center', cellLoc='center',
                       colWidths=[0.12, 0.30, 0.12, 0.12, 0.15, 0.15])
    table2.auto_set_font_size(False)
    table2.set_fontsize(9)
    table2.scale(1.2, 1.5)
    
    for j in range(6):
        table2[(0, j)].set_facecolor('#2E86AB')
        table2[(0, j)].set_text_props(color='white', fontweight='bold')
    
    ax4.set_title('Peer Company Details', fontsize=12, fontweight='bold', pad=20)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"  Chart saved: {output_path}")


def create_excel_output(allegion_data, allegion_results, peer_results, peer_stats, output_path):
    """Create Excel workbook with all analysis."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not available. Skipping Excel output.")
        return
    
    wb = Workbook()
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # ===== Sheet 1: Summary =====
    ws1 = wb.active
    ws1.title = "Summary"
    
    ws1['A1'] = "ALLEGION (ALLE) BETA ANALYSIS WITH PEER COMPARISON"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A2'] = f"Data Source: 2024 Form 10-K (Filed {allegion_data['filing_date']})"
    
    summary = [
        ["", "", ""],
        ["ALLEGION REGRESSION BETA", "", ""],
        ["Levered Beta (Equity)", allegion_results['levered_beta'], "From 5yr monthly regression"],
        ["Unlevered Beta (Asset)", allegion_results['unlevered_beta'], "Hamada equation"],
        ["R-Squared", allegion_results['r_squared'], ""],
        ["Observations", allegion_results['n_observations'], "Monthly returns"],
        ["", "", ""],
        ["PEER ANALYSIS", "", ""],
        ["Number of Peers", len(peer_results), ""],
        ["Peer Avg Unlevered Beta", peer_stats['avg_unlevered'], ""],
        ["Peer Median Unlevered Beta", peer_stats['median_unlevered'], ""],
        ["Peer Std Dev", peer_stats['std_unlevered'], ""],
        ["", "", ""],
        ["RE-LEVERED BETA (using ALLE capital structure)", "", ""],
        ["From Peer Average", peer_stats['relevered_from_avg'], f"D/E={allegion_data['debt_to_equity']:.4f}"],
        ["From Peer Median", peer_stats['relevered_from_median'], f"Tax={allegion_data['effective_tax_rate']*100:.1f}%"],
        ["", "", ""],
        ["RECOMMENDED BETA", "", ""],
        ["Average of Methods", peer_stats['recommended_beta'], "Avg of regression & peer re-levered"],
    ]
    
    for r, row in enumerate(summary, 4):
        for c, val in enumerate(row, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if isinstance(val, float):
                cell.number_format = '0.0000'
            if r in [5, 11, 17, 21]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.font = header_font
            if r == 22:
                cell.fill = highlight_fill
                cell.font = Font(bold=True)
    
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 30
    
    # ===== Sheet 2: Peer Details =====
    ws2 = wb.create_sheet("Peer Analysis")
    
    ws2['A1'] = "PEER COMPANY BETA ANALYSIS"
    ws2['A1'].font = Font(bold=True, size=14)
    
    peer_headers = ["Ticker", "Company Name", "Total Debt ($M)", "Total Equity ($M)", 
                    "D/E Ratio", "Tax Rate", "Levered Beta", "Unlevered Beta"]
    
    for c, header in enumerate(peer_headers, 1):
        cell = ws2.cell(row=3, column=c, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    for r, peer in enumerate(peer_results, 4):
        data_row = [
            peer['ticker'],
            peer['name'],
            peer['total_debt'],
            peer['total_equity'],
            peer['d_e_ratio'],
            peer['tax_rate'],
            peer['levered_beta'],
            peer['unlevered_beta']
        ]
        for c, val in enumerate(data_row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if isinstance(val, float):
                if c in [3, 4]:
                    cell.number_format = '#,##0.0'
                elif c == 6:
                    cell.number_format = '0.0%'
                else:
                    cell.number_format = '0.0000'
    
    # Add averages
    avg_row = len(peer_results) + 5
    ws2.cell(row=avg_row, column=1, value="AVERAGE").font = Font(bold=True)
    ws2.cell(row=avg_row, column=5, value=np.mean([p['d_e_ratio'] for p in peer_results])).number_format = '0.0000'
    ws2.cell(row=avg_row, column=7, value=peer_stats['avg_levered']).number_format = '0.0000'
    ws2.cell(row=avg_row, column=8, value=peer_stats['avg_unlevered']).number_format = '0.0000'
    
    ws2.cell(row=avg_row+1, column=1, value="MEDIAN").font = Font(bold=True)
    ws2.cell(row=avg_row+1, column=5, value=np.median([p['d_e_ratio'] for p in peer_results])).number_format = '0.0000'
    ws2.cell(row=avg_row+1, column=7, value=peer_stats['median_levered']).number_format = '0.0000'
    ws2.cell(row=avg_row+1, column=8, value=peer_stats['median_unlevered']).number_format = '0.0000'
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws2.column_dimensions[col].width = 18
    ws2.column_dimensions['B'].width = 28
    
    # ===== Sheet 3: Allegion Data =====
    ws3 = wb.create_sheet("Allegion 10-K Data")
    
    ws3['A1'] = "ALLEGION FINANCIAL DATA FROM 2024 10-K"
    ws3['A1'].font = Font(bold=True, size=14)
    
    alle_data = [
        ["Item", "Value", "10-K Reference"],
        ["Total Debt", allegion_data['total_debt'], "Page F-4"],
        ["Total Equity", allegion_data['total_equity'], "Page F-4"],
        ["Cash & Equivalents", allegion_data['cash'], "Page F-4"],
        ["Debt-to-Equity Ratio", allegion_data['debt_to_equity'], "Calculated"],
        ["Effective Tax Rate", allegion_data['effective_tax_rate'], "Note 18, Page F-27"],
        ["Operating Income", allegion_data['operating_income'], "Page F-3"],
        ["Interest Expense", allegion_data['interest_expense'], "Page F-3"],
        ["Interest Coverage", allegion_data['interest_coverage'], "Calculated"],
    ]
    
    for r, row in enumerate(alle_data, 3):
        for c, val in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if r == 3:
                cell.fill = header_fill
                cell.font = header_font
            if isinstance(val, float):
                if c == 2 and r in [4, 5, 6, 9, 10]:
                    cell.number_format = '#,##0.0'
                elif c == 2 and r == 8:
                    cell.number_format = '0.0000'
                elif c == 2 and r == 9:
                    cell.number_format = '0.0%'
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 25
    
    # ===== Sheet 4: Calculations =====
    ws4 = wb.create_sheet("Calculations")
    
    calcs = [
        ["BETA CALCULATION FORMULAS AND STEPS", ""],
        ["", ""],
        ["1. RAW BETA (Regression)", ""],
        ["Formula", "β = Cov(R_stock, R_market) / Var(R_market)"],
        ["Covariance", allegion_results['covariance']],
        ["Market Variance", allegion_results['market_variance']],
        ["Calculated Beta", allegion_results['levered_beta']],
        ["", ""],
        ["2. UNLEVER BETA (Hamada Equation)", ""],
        ["Formula", "β_U = β_L / [1 + (1-T) × D/E]"],
        ["Levered Beta", allegion_results['levered_beta']],
        ["D/E Ratio", allegion_data['debt_to_equity']],
        ["Tax Rate", allegion_data['effective_tax_rate']],
        ["Leverage Factor", 1 + (1 - allegion_data['effective_tax_rate']) * allegion_data['debt_to_equity']],
        ["Unlevered Beta", allegion_results['unlevered_beta']],
        ["", ""],
        ["3. PEER AVERAGE APPROACH", ""],
        ["Peer Avg Unlevered Beta", peer_stats['avg_unlevered']],
        ["Re-lever Formula", "β_L = β_U × [1 + (1-T) × D/E]"],
        ["Re-levered Beta (from avg)", peer_stats['relevered_from_avg']],
        ["", ""],
        ["4. PEER MEDIAN APPROACH", ""],
        ["Peer Median Unlevered Beta", peer_stats['median_unlevered']],
        ["Re-levered Beta (from median)", peer_stats['relevered_from_median']],
    ]
    
    for r, row in enumerate(calcs, 1):
        for c, val in enumerate(row, 1):
            cell = ws4.cell(row=r, column=c, value=val)
            if r in [1, 3, 9, 17, 22]:
                cell.font = Font(bold=True)
            if isinstance(val, float):
                cell.number_format = '0.000000' if val < 0.01 else '0.0000'
    
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 45
    
    wb.save(output_path)
    print(f"  Excel saved: {output_path}")


def print_header(text, char='='):
    """Print formatted section header."""
    print(f"\n{char * 70}")
    print(f" {text}")
    print(f"{char * 70}")


# ============================================================================
# MAIN ANALYSIS
# ============================================================================

def main():
    """Run complete beta analysis."""
    
    print("\n" + "=" * 70)
    print("   ALLEGION (ALLE) BETA ANALYSIS WITH PEER COMPARISON")
    print(f"   Using {LOOKBACK_YEARS}-year {RETURN_PERIOD} returns")
    print("=" * 70)
    
    # =========================================================================
    # STEP 1: Display Allegion Financial Data
    # =========================================================================
    print_header("STEP 1: ALLEGION FINANCIAL DATA (2024 10-K)")
    
    print(f"""
Source: Allegion 2024 Form 10-K (Filed {ALLEGION_DATA['filing_date']})

Capital Structure (Page F-4):
  Total Debt:               ${ALLEGION_DATA['total_debt']:,.1f}M
  Total Equity:             ${ALLEGION_DATA['total_equity']:,.1f}M
  D/E Ratio:                {ALLEGION_DATA['debt_to_equity']:.4f}
  Cash:                     ${ALLEGION_DATA['cash']:,.1f}M
  Net Debt:                 ${ALLEGION_DATA['net_debt']:,.1f}M

Tax Rate (Note 18, Page F-27):
  Effective Tax Rate:       {ALLEGION_DATA['effective_tax_rate']*100:.1f}%

Interest Coverage:
  Operating Income:         ${ALLEGION_DATA['operating_income']:,.1f}M
  Interest Expense:         ${ALLEGION_DATA['interest_expense']:,.1f}M
  Coverage Ratio:           {ALLEGION_DATA['interest_coverage']:.2f}x
    """)
    
    # =========================================================================
    # STEP 2: Calculate Allegion's Regression Beta
    # =========================================================================
    print_header("STEP 2: ALLEGION REGRESSION BETA")
    
    print(f"\nFetching {LOOKBACK_YEARS}-year {RETURN_PERIOD} price data for ALLE and S&P 500...")
    
    # Fetch data
    alle_data, market_data = fetch_stock_data('ALLE', '^GSPC', years=LOOKBACK_YEARS)
    
    if alle_data is not None and market_data is not None:
        alle_returns = calculate_returns(alle_data, RETURN_PERIOD)
        market_returns = calculate_returns(market_data, RETURN_PERIOD)
        print(f"  Retrieved {len(alle_returns)} {RETURN_PERIOD} observations")
    else:
        print("  Using simulated data for demonstration...")
        alle_returns, market_returns = generate_sample_returns()
    
    # Calculate regression beta
    alle_regression = calculate_beta_regression(alle_returns, market_returns)
    
    alle_levered_beta = alle_regression['beta']
    alle_unlevered_beta = unlever_beta(alle_levered_beta, 
                                        ALLEGION_DATA['debt_to_equity'],
                                        ALLEGION_DATA['effective_tax_rate'])
    
    leverage_factor = 1 + (1 - ALLEGION_DATA['effective_tax_rate']) * ALLEGION_DATA['debt_to_equity']
    
    print(f"""
Regression Results ({LOOKBACK_YEARS}-year {RETURN_PERIOD} returns):
  Raw (Levered) Beta:       {alle_levered_beta:.4f}
  R-squared:                {alle_regression['r_squared']:.4f}
  Correlation:              {alle_regression['correlation']:.4f}
  Number of observations:   {alle_regression['n_observations']}

Unlevering Calculation (Hamada Equation):
  Formula: β_U = β_L / [1 + (1-T) × D/E]
  
  β_U = {alle_levered_beta:.4f} / [1 + (1 - {ALLEGION_DATA['effective_tax_rate']:.3f}) × {ALLEGION_DATA['debt_to_equity']:.4f}]
  β_U = {alle_levered_beta:.4f} / {leverage_factor:.4f}
  β_U = {alle_unlevered_beta:.4f}
    """)
    
    # Store Allegion results
    allegion_results = {
        'levered_beta': alle_levered_beta,
        'unlevered_beta': alle_unlevered_beta,
        'r_squared': alle_regression['r_squared'],
        'correlation': alle_regression['correlation'],
        'n_observations': alle_regression['n_observations'],
        'covariance': alle_regression['covariance'],
        'market_variance': alle_regression['market_variance'],
    }
    
    # =========================================================================
    # STEP 3: Calculate Peer Betas
    # =========================================================================
    print_header("STEP 3: PEER COMPANY BETA ANALYSIS")
    
    print("\nFetching data for peer companies...")
    print(f"{'Ticker':<8} {'Company':<28} {'D/E':<8} {'Tax':<8} {'Lev β':<10} {'Unlev β':<10}")
    print("-" * 80)
    
    # First fetch market data once
    _, market_data_peers = fetch_stock_data('^GSPC', '^GSPC', years=LOOKBACK_YEARS)
    if market_data_peers is not None:
        market_returns_peers = calculate_returns(market_data_peers, RETURN_PERIOD)
    else:
        _, market_returns_peers = generate_sample_returns()
    
    peer_results = []
    
    for ticker, peer_info in PEER_COMPANIES.items():
        if not peer_info['include']:
            continue
        
        # Fetch peer stock data
        peer_stock_data, _ = fetch_stock_data(ticker, '^GSPC', years=LOOKBACK_YEARS)
        
        if peer_stock_data is not None and len(peer_stock_data) > 0:
            peer_returns = calculate_returns(peer_stock_data, RETURN_PERIOD)
            peer_regression = calculate_beta_regression(peer_returns, market_returns_peers)
            
            if peer_regression is not None:
                levered_beta = peer_regression['beta']
            else:
                # Use estimated beta based on industry
                levered_beta = 1.0 + 0.3 * (peer_info['debt_to_equity'] - 1.0)
        else:
            # Estimate beta if data unavailable
            levered_beta = 1.0 + 0.3 * (peer_info['debt_to_equity'] - 1.0)
        
        # Calculate unlevered beta
        unlevered_beta = unlever_beta(levered_beta, peer_info['debt_to_equity'], 
                                       peer_info['effective_tax_rate'])
        
        peer_result = {
            'ticker': ticker,
            'name': peer_info['name'],
            'total_debt': peer_info['total_debt'],
            'total_equity': peer_info['total_equity'],
            'd_e_ratio': peer_info['debt_to_equity'],
            'tax_rate': peer_info['effective_tax_rate'],
            'levered_beta': levered_beta,
            'unlevered_beta': unlevered_beta,
        }
        peer_results.append(peer_result)
        
        print(f"{ticker:<8} {peer_info['name'][:26]:<28} {peer_info['debt_to_equity']:<8.2f} "
              f"{peer_info['effective_tax_rate']*100:<7.0f}% {levered_beta:<10.4f} {unlevered_beta:<10.4f}")
    
    # =========================================================================
    # STEP 4: Calculate Peer Statistics
    # =========================================================================
    print_header("STEP 4: PEER BETA STATISTICS")
    
    levered_betas = [p['levered_beta'] for p in peer_results]
    unlevered_betas = [p['unlevered_beta'] for p in peer_results]
    
    avg_levered = np.mean(levered_betas)
    median_levered = np.median(levered_betas)
    avg_unlevered = np.mean(unlevered_betas)
    median_unlevered = np.median(unlevered_betas)
    std_unlevered = np.std(unlevered_betas)
    
    print(f"""
Peer Levered Beta:
  Average:                 {avg_levered:.4f}
  Median:                  {median_levered:.4f}

Peer Unlevered Beta:
  Average:                 {avg_unlevered:.4f}
  Median:                  {median_unlevered:.4f}
  Std Deviation:           {std_unlevered:.4f}
  Range:                   [{min(unlevered_betas):.4f}, {max(unlevered_betas):.4f}]
    """)
    
    # =========================================================================
    # STEP 5: Re-lever Using Allegion's Capital Structure
    # =========================================================================
    print_header("STEP 5: RE-LEVER PEER BETA FOR ALLEGION")
    
    relevered_from_avg = relever_beta(avg_unlevered, ALLEGION_DATA['debt_to_equity'], 
                                       ALLEGION_DATA['effective_tax_rate'])
    relevered_from_median = relever_beta(median_unlevered, ALLEGION_DATA['debt_to_equity'], 
                                          ALLEGION_DATA['effective_tax_rate'])
    
    leverage_factor = 1 + (1 - ALLEGION_DATA['effective_tax_rate']) * ALLEGION_DATA['debt_to_equity']
    
    print(f"""
Re-levering Formula: β_L = β_U × [1 + (1-T) × D/E]

Using Allegion's Capital Structure:
  D/E Ratio:               {ALLEGION_DATA['debt_to_equity']:.4f}
  Tax Rate:                {ALLEGION_DATA['effective_tax_rate']*100:.1f}%
  Leverage Factor:         {leverage_factor:.4f}

From Peer Average Unlevered Beta ({avg_unlevered:.4f}):
  Re-levered β = {avg_unlevered:.4f} × {leverage_factor:.4f} = {relevered_from_avg:.4f}

From Peer Median Unlevered Beta ({median_unlevered:.4f}):
  Re-levered β = {median_unlevered:.4f} × {leverage_factor:.4f} = {relevered_from_median:.4f}
    """)
    
    # Store peer statistics
    peer_stats = {
        'avg_levered': avg_levered,
        'median_levered': median_levered,
        'avg_unlevered': avg_unlevered,
        'median_unlevered': median_unlevered,
        'std_unlevered': std_unlevered,
        'relevered_from_avg': relevered_from_avg,
        'relevered_from_median': relevered_from_median,
        'recommended_beta': np.mean([alle_levered_beta, relevered_from_avg, relevered_from_median]),
    }
    
    # =========================================================================
    # STEP 6: Final Summary
    # =========================================================================
    print_header("FINAL SUMMARY")
    
    print(f"""
┌──────────────────────────────────────────────────────────────────────────┐
│                    ALLEGION (ALLE) BETA ANALYSIS                         │
├──────────────────────────────────────────────────────────────────────────┤
│                                                                          │
│  METHOD 1: REGRESSION ({LOOKBACK_YEARS}-year {RETURN_PERIOD} returns)                             │
│    • Levered Beta:                    {alle_levered_beta:>8.4f}                         │
│    • Unlevered Beta:                  {alle_unlevered_beta:>8.4f}                         │
│    • R-squared:                       {alle_regression['r_squared']:>8.4f}                         │
│    • Observations:                    {alle_regression['n_observations']:>8}                         │
│                                                                          │
│  METHOD 2: PEER ANALYSIS ({len(peer_results)} companies)                                │
│    • Peer Avg Unlevered Beta:         {avg_unlevered:>8.4f}                         │
│    • Peer Median Unlevered Beta:      {median_unlevered:>8.4f}                         │
│    • Re-levered (from avg):           {relevered_from_avg:>8.4f}                         │
│    • Re-levered (from median):        {relevered_from_median:>8.4f}                         │
│                                                                          │
│  ALLEGION CAPITAL STRUCTURE (10-K):                                      │
│    • Total Debt:                   ${ALLEGION_DATA['total_debt']:>8,.1f}M                        │
│    • Total Equity:                 ${ALLEGION_DATA['total_equity']:>8,.1f}M                        │
│    • D/E Ratio:                       {ALLEGION_DATA['debt_to_equity']:>8.4f}                         │
│    • Effective Tax Rate:              {ALLEGION_DATA['effective_tax_rate']*100:>7.1f}%                         │
│                                                                          │
├──────────────────────────────────────────────────────────────────────────┤
│  RECOMMENDED BETA (average of methods):    {peer_stats['recommended_beta']:>8.4f}                    │
└──────────────────────────────────────────────────────────────────────────┘
    """)
    
    # =========================================================================
    # STEP 7: Generate Output Files
    # =========================================================================
    print_header("GENERATING OUTPUT FILES")
    
    output_dir = '.'
    
    # Regression chart for Allegion
    create_regression_chart(
        alle_regression['stock_returns'], 
        alle_regression['market_returns'],
        alle_regression,
        company_name='ALLE',
        output_path=f'{output_dir}/allegion_regression_chart_weekly.jpeg'
    )
    
    # Peer comparison chart
    allegion_beta_dict = {'levered': alle_levered_beta, 'unlevered': alle_unlevered_beta}
    create_peer_comparison_chart(peer_results, allegion_beta_dict, 
                                  f'{output_dir}/peer_beta_comparison_weekly.jpeg')
    
    # Summary dashboard
    create_summary_dashboard(ALLEGION_DATA, allegion_results, peer_results, peer_stats,
                             f'{output_dir}/beta_analysis_dashboard_weekly.jpeg')
    
    # Excel output
    create_excel_output(ALLEGION_DATA, allegion_results, peer_results, peer_stats,
                        f'{output_dir}/allegion_beta_peer_analysis_weekly.xlsx')
    
    print("\n" + "=" * 70)
    print(" ANALYSIS COMPLETE")
    print("=" * 70)
    
    return {
        'allegion': allegion_results,
        'peers': peer_results,
        'peer_stats': peer_stats,
    }


if __name__ == "__main__":
    results = main()
